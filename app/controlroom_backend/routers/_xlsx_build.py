"""Build openpyxl workbooks for export and template generation."""
import io
import json
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from routers._xlsx_schema import (
    ColDef, TableConfig, TABLE_CONFIGS,
    LOOKUP_COLS, LOOKUP_MAX_ROWS,
)

_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_ID_FILL = PatternFill("solid", fgColor="374151")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_COL_WIDTHS = {"id_only": 38, "multi": 40, "default": 22}


async def fetch_lookup_data(conn) -> dict[str, list[tuple[UUID, str]]]:
    result: dict[str, list[tuple[UUID, str]]] = {}
    brands = await conn.fetch(
        "SELECT brand_id, brand_name, legal_name FROM brands "
        "WHERE deleted_at IS NULL ORDER BY COALESCE(brand_name, legal_name)"
    )
    result["brands"] = [
        (r["brand_id"], r["brand_name"] or r["legal_name"] or "")
        for r in brands if r["brand_name"] or r["legal_name"]
    ]
    for key, table in [
        ("entity_types",    "entity_types"),
        ("plugin_formats",  "plugin_formats"),
        ("effect_types",    "effect_types"),
        ("instrument_types","instrument_types"),
        ("tool_types",      "tool_types"),
        ("model_types",     "model_types"),
        ("tag_types",       "tag_types"),
    ]:
        rows = await conn.fetch(
            f"SELECT type_id, type_name FROM {table} WHERE deleted_at IS NULL ORDER BY type_name"
        )
        result[key] = [(r["type_id"], r["type_name"]) for r in rows]
    return result


def _cell_value(row: dict, col: ColDef) -> Any:
    val = row.get(col.export_key)
    if val is None:
        return None
    if col.id_only:
        return str(val)
    if col.multi and isinstance(val, list):
        return ", ".join(item["name"] for item in val if item.get("name"))
    if col.export_key == "attributes" and isinstance(val, dict):
        return json.dumps(val)
    return val


def _write_lookup_sheet(wb: Workbook, lookup_data: dict) -> dict[str, str]:
    ws = wb.create_sheet("_Lookups")
    ws.sheet_state = "hidden"
    ranges: dict[str, str] = {}
    for col_idx, key in enumerate(LOOKUP_COLS, start=1):
        letter = get_column_letter(col_idx)
        ws.cell(row=1, column=col_idx, value=key)
        for row_idx, (_, name) in enumerate(lookup_data.get(key, []), start=2):
            ws.cell(row=row_idx, column=col_idx, value=name)
        ranges[key] = f"'_Lookups'!${letter}$2:${letter}${LOOKUP_MAX_ROWS}"
    return ranges


def _write_header(ws, visible_cols: list[ColDef]) -> None:
    for i, col in enumerate(visible_cols, start=1):
        cell = ws.cell(row=1, column=i, value=col.header)
        cell.font = _HEADER_FONT
        cell.fill = _ID_FILL if col.id_only else _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        if col.id_only:
            width = _COL_WIDTHS["id_only"]
        elif col.multi:
            width = _COL_WIDTHS["multi"]
        else:
            width = _COL_WIDTHS["default"]
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20


def _add_validations(ws, visible_cols: list[ColDef], ranges: dict, max_rows: int = 5000) -> None:
    for i, col in enumerate(visible_cols, start=1):
        if not col.lookup or col.multi or col.lookup not in ranges:
            continue
        col_letter = get_column_letter(i)
        dv = DataValidation(
            type="list",
            formula1=ranges[col.lookup],
            showErrorMessage=True,
            errorTitle="Value not in list",
            error="This value was not found. The import will fail validation.",
            errorStyle="warning",
        )
        ws.add_data_validation(dv)
        dv.sqref = f"{col_letter}2:{col_letter}{max_rows}"


def _write_data_rows(ws, visible_cols: list[ColDef], rows: list[dict]) -> None:
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(visible_cols, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_cell_value(row, col))


async def fetch_table_rows(conn, config: TableConfig) -> list[dict]:
    rows = await conn.fetch(f"SELECT * FROM {config.view}")
    return [dict(r) for r in rows]


def build_workbook(
    table_keys: list[str],
    rows_by_key: dict[str, list[dict]],
    lookup_data: dict,
    is_template: bool,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    lookup_ranges = _write_lookup_sheet(wb, lookup_data)
    for key in table_keys:
        config = TABLE_CONFIGS[key]
        visible_cols = [c for c in config.cols if not (c.id_only and is_template)]
        ws = wb.create_sheet(config.sheet_name)
        _write_header(ws, visible_cols)
        if not is_template:
            _write_data_rows(ws, visible_cols, rows_by_key.get(key, []))
        _add_validations(ws, visible_cols, lookup_ranges)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
