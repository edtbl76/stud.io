"""Parse xlsx files and execute validated imports."""
import difflib
import io
import json
from typing import Any
from uuid import UUID

from asyncpg import Connection
from openpyxl import load_workbook

from routers._helpers import log_audit, _serializable
from routers._xlsx_schema import TableConfig, TABLE_CONFIGS, SHEET_TO_KEY


def parse_workbook(file_bytes: bytes) -> dict[str, list[dict]]:
    """Return {sheet_name: [{header: cell_value, ...}, ...]} skipping _Lookups."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    result: dict[str, list[dict]] = {}
    for name in wb.sheetnames:
        if name.startswith("_"):
            continue
        ws = wb[name]
        all_rows = list(ws.rows)
        if not all_rows:
            continue
        headers = [c.value for c in all_rows[0]]
        result[name] = [
            {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
            for row in all_rows[1:]
            if any(cell.value is not None for cell in row)
        ]
    return result


def _name_map(lookup_data: dict, key: str) -> dict[str, UUID]:
    return {name.lower(): uid for uid, name in lookup_data.get(key, [])}


def _fuzzy(val: str, candidates: list[str]) -> str | None:
    matches = difflib.get_close_matches(val.lower(), [c.lower() for c in candidates], n=1, cutoff=0.6)
    if not matches:
        return None
    idx = [c.lower() for c in candidates].index(matches[0])
    return candidates[idx]


def _err(sheet: str, row: int, col: str, val: str, close: str | None) -> dict:
    msg = f"'{val}' not found"
    if close:
        msg += f" — did you mean '{close}'?"
    return {"sheet": sheet, "row": row, "column": col, "value": val, "message": msg}


def _resolve_col(col_def, raw_val: str, lookup_data: dict, sheet: str, row: int) -> tuple[Any, list[dict]]:
    nmap = _name_map(lookup_data, col_def.lookup)
    names = list(nmap.keys())
    if not col_def.multi:
        resolved = nmap.get(raw_val.lower())
        if resolved is None:
            return None, [_err(sheet, row, col_def.header, raw_val, _fuzzy(raw_val, names))]
        return resolved, []
    parts = [p.strip() for p in raw_val.split(",") if p.strip()]
    ids, errors = [], []
    for part in parts:
        resolved = nmap.get(part.lower())
        if resolved is None:
            errors.append(_err(sheet, row, col_def.header, part, _fuzzy(part, names)))
        else:
            ids.append(resolved)
    return (ids if ids else None), errors


def _prep_col(col, raw: Any, lookup_data: dict, sheet: str, row_num: int) -> tuple[Any, list[dict]]:
    """Prepare the value for a single column. Returns (value, errors)."""
    str_val = str(raw).strip() if raw is not None else ""
    if not str_val:
        return None, []
    if col.lookup:
        return _resolve_col(col, str_val, lookup_data, sheet, row_num)
    if col.export_key == "attributes":
        try:
            return json.dumps(json.loads(str_val)), []
        except ValueError:
            err = _err(sheet, row_num, col.header, str_val[:40], None)
            return None, [{**err, "message": "Invalid JSON"}]
    return str_val, []


def _prep_row(row_dict: dict, config: TableConfig, lookup_data: dict, sheet: str, row_num: int) -> tuple[dict, list[dict]]:
    data: dict[str, Any] = {}
    errors: list[dict] = []
    for col in config.cols:
        if col.id_only or col.write_key is None:
            continue
        value, errs = _prep_col(col, row_dict.get(col.header), lookup_data, sheet, row_num)
        data[col.write_key] = value
        errors.extend(errs)
    return data, errors


def _validate_id(sheet: str, row: int, id_val: str) -> dict | None:
    try:
        UUID(id_val)
        return None
    except ValueError:
        return _err(sheet, row, "ID", id_val, None) | {"message": "Invalid UUID"}


def validate_import(parsed: dict[str, list[dict]], lookup_data: dict) -> list[dict]:
    errors: list[dict] = []
    for sheet_name, rows in parsed.items():
        key = SHEET_TO_KEY.get(sheet_name)
        if not key:
            continue
        config = TABLE_CONFIGS[key]
        for i, row in enumerate(rows, start=2):
            id_val = row.get("ID")
            if id_val is not None:
                id_err = _validate_id(sheet_name, i, str(id_val))
                if id_err:
                    errors.append(id_err)
            _, row_errors = _prep_row(row, config, lookup_data, sheet_name, i)
            errors.extend(row_errors)
    return errors


async def _write_row(conn: Connection, config: TableConfig, data: dict, record_id: UUID | None) -> UUID:
    write_cols = [c for c in config.cols if not c.id_only and c.write_key]
    col_names = [c.write_key for c in write_cols]
    values = [data.get(c.write_key) for c in write_cols]
    if record_id is not None:
        set_clause = ", ".join(f"{n}=${i + 2}" for i, n in enumerate(col_names))
        sql = (
            f"UPDATE {config.table} SET {set_clause} "
            f"WHERE {config.id_col}=$1 AND deleted_at IS NULL "
            f"RETURNING {config.id_col}"
        )
        row = await conn.fetchrow(sql, record_id, *values)
        if not row:
            raise ValueError(f"ID {record_id} not found or deleted in {config.table}")
        return row[config.id_col]
    placeholders = ", ".join(f"${i + 1}" for i in range(len(col_names)))
    sql = (
        f"INSERT INTO {config.table} ({', '.join(col_names)}) "
        f"VALUES ({placeholders}) RETURNING {config.id_col}"
    )
    row = await conn.fetchrow(sql, *values)
    return row[config.id_col]


async def _import_row(
    conn: Connection, config: TableConfig, row: dict, lookup_data: dict,
    sheet: str, row_num: int, performed_by: str,
) -> str:
    """Write one row and record an audit entry. Returns 'create' or 'update'."""
    data, _ = _prep_row(row, config, lookup_data, sheet, row_num)
    id_val = row.get("ID")
    record_id = UUID(str(id_val)) if id_val else None
    async with conn.transaction():
        old_row = await conn.fetchrow(
            f"SELECT * FROM {config.table} WHERE {config.id_col} = $1 FOR UPDATE",  # safe: config constants
            record_id,
        ) if record_id is not None else None
        new_id = await _write_row(conn, config, data, record_id)
        new_row = await conn.fetchrow(
            f"SELECT * FROM {config.table} WHERE {config.id_col} = $1",  # safe: config constants
            new_id,
        )
        await log_audit(
            conn, config.table, new_id,
            "UPDATE" if record_id else "CREATE",
            performed_by=performed_by,
            old_data=_serializable(dict(old_row)) if old_row else None,
            new_data=_serializable(dict(new_row)) if new_row else None,
        )
    return "update" if record_id else "create"


async def execute_import(conn: Connection, parsed: dict[str, list[dict]], lookup_data: dict, performed_by: str) -> list[dict]:
    summary: list[dict] = []
    for sheet_name, rows in parsed.items():
        key = SHEET_TO_KEY.get(sheet_name)
        if not key:
            continue
        config = TABLE_CONFIGS[key]
        creates = updates = 0
        for i, row in enumerate(rows, start=2):
            result = await _import_row(conn, config, row, lookup_data, sheet_name, i, performed_by)
            if result == "update":
                updates += 1
            else:
                creates += 1
        summary.append({"sheet": sheet_name, "creates": creates, "updates": updates})
    return summary
