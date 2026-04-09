"""Tests for import/export endpoints."""
import io

import openpyxl


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

async def test_export_xlsx_returns_xlsx(client, admin_headers):
    res = await client.get("/admin/export/xlsx?tables=brands", headers=admin_headers)
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]
    assert res.headers["content-disposition"].endswith(".xlsx")


async def test_export_xlsx_contains_correct_sheets(client, admin_headers):
    res = await client.get("/admin/export/xlsx?tables=brands,effects", headers=admin_headers)
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert "Brands" in wb.sheetnames
    assert "Effects" in wb.sheetnames


async def test_export_xlsx_header_row(client, admin_headers):
    res = await client.get("/admin/export/xlsx?tables=brands", headers=admin_headers)
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb["Brands"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, 10) if ws.cell(row=1, column=i).value]
    assert "ID" in headers
    assert "Name" in headers
    assert "Legal Name" in headers


async def test_export_xlsx_has_data_rows(client, admin_headers):
    res = await client.get("/admin/export/xlsx?tables=brands", headers=admin_headers)
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb["Brands"]
    assert ws.max_row > 1


async def test_export_xlsx_requires_admin(client, auth_headers):
    res = await client.get("/admin/export/xlsx?tables=brands", headers=auth_headers)
    assert res.status_code == 403


async def test_export_xlsx_unknown_table(client, admin_headers):
    res = await client.get("/admin/export/xlsx?tables=nonexistent", headers=admin_headers)
    assert res.status_code == 400


async def test_export_xlsx_no_tables(client, admin_headers):
    res = await client.get("/admin/export/xlsx?tables=", headers=admin_headers)
    assert res.status_code == 400


# ---------------------------------------------------------------------------
# Template
# ---------------------------------------------------------------------------

async def test_template_returns_xlsx(client, admin_headers):
    res = await client.get("/admin/export/template?tables=effects", headers=admin_headers)
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]


async def test_template_has_no_data_rows(client, admin_headers):
    res = await client.get("/admin/export/template?tables=effects", headers=admin_headers)
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb["Effects"]
    assert ws.max_row == 1


async def test_template_omits_id_column(client, admin_headers):
    res = await client.get("/admin/export/template?tables=brands", headers=admin_headers)
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb["Brands"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, 15) if ws.cell(row=1, column=i).value]
    assert "ID" not in headers
    assert "Name" in headers


async def test_template_has_lookup_sheet(client, admin_headers):
    res = await client.get("/admin/export/template?tables=brands", headers=admin_headers)
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert "_Lookups" in wb.sheetnames


# ---------------------------------------------------------------------------
# Import — validation errors
# ---------------------------------------------------------------------------

def _make_xlsx(sheets: dict[str, list[dict]], headers_by_sheet: dict[str, list[str]] | None = None) -> bytes:
    """Build an xlsx bytes fixture from {sheet_name: [row_dicts]}."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        extra_headers = (headers_by_sheet or {}).get(sheet_name)
        headers = extra_headers if extra_headers else (list(rows[0].keys()) if rows else [])
        for c_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=c_idx, value=h)
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, h in enumerate(headers, start=1):
                ws.cell(row=r_idx, column=c_idx, value=row.get(h))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_import_invalid_brand_name_returns_422(client, admin_headers):
    data = _make_xlsx({"Effects": [{"Name": "Test", "Brand": "ZZZ_NO_MATCH_ZZZ"}]})
    res = await client.post(
        "/admin/import/xlsx",
        headers=admin_headers,
        files={"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 422
    detail = res.json()["detail"]
    assert "errors" in detail
    assert any(e["column"] == "Brand" for e in detail["errors"])


async def test_import_invalid_id_returns_422(client, admin_headers):
    data = _make_xlsx({"Brands": [{"ID": "not-a-uuid", "Name": "Bad ID Brand"}]})
    res = await client.post(
        "/admin/import/xlsx",
        headers=admin_headers,
        files={"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 422
    detail = res.json()["detail"]
    assert any(e["column"] == "ID" for e in detail["errors"])


async def test_import_wrong_extension_returns_400(client, admin_headers):
    res = await client.post(
        "/admin/import/xlsx",
        headers=admin_headers,
        files={"file": ("test.csv", b"a,b,c", "text/csv")},
    )
    assert res.status_code == 400


async def test_import_requires_admin(client, auth_headers):
    data = _make_xlsx({"Brands": []})
    res = await client.post(
        "/admin/import/xlsx",
        headers=auth_headers,
        files={"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 403


# ---------------------------------------------------------------------------
# Import — successful create
# ---------------------------------------------------------------------------

async def test_import_creates_brand(client, admin_headers, conn):
    data = _make_xlsx({"Brands": [{"Name": "Import Test Brand", "Legal Name": "Import Test Legal"}]})
    res = await client.post(
        "/admin/import/xlsx",
        headers=admin_headers,
        files={"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    result = res.json()
    assert result["total_creates"] == 1
    assert result["total_updates"] == 0
    row = await conn.fetchrow("SELECT brand_name FROM brands WHERE brand_name = 'Import Test Brand'")
    assert row is not None


async def test_import_updates_brand(client, admin_headers, conn):
    brand_id = await conn.fetchval(
        "INSERT INTO brands (brand_name) VALUES ('UpdateMe') RETURNING brand_id"
    )
    data = _make_xlsx({"Brands": [{"ID": str(brand_id), "Name": "UpdateMe Updated"}]})
    res = await client.post(
        "/admin/import/xlsx",
        headers=admin_headers,
        files={"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    result = res.json()
    assert result["total_updates"] == 1
    row = await conn.fetchrow("SELECT brand_name FROM brands WHERE brand_id = $1", brand_id)
    assert row["brand_name"] == "UpdateMe Updated"


async def test_import_empty_sheet_ignored(client, admin_headers):
    data = _make_xlsx({"Brands": []}, headers_by_sheet={"Brands": ["Name", "Legal Name"]})
    res = await client.post(
        "/admin/import/xlsx",
        headers=admin_headers,
        files={"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    assert res.json()["total_creates"] == 0


async def test_import_unknown_sheet_ignored(client, admin_headers):
    data = _make_xlsx({"UnknownTable": [{"Name": "Test"}]})
    res = await client.post(
        "/admin/import/xlsx",
        headers=admin_headers,
        files={"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200


# ---------------------------------------------------------------------------
# Import — audit log
# ---------------------------------------------------------------------------

async def test_import_create_writes_audit_entry(client, admin_headers, conn):
    data = _make_xlsx({"Brands": [{"Name": "Audit Create Brand", "Legal Name": "ACB Legal"}]})
    res = await client.post(
        "/admin/import/xlsx",
        headers=admin_headers,
        files={"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200

    row = await conn.fetchrow(
        "SELECT operation, old_data, new_data, performed_by FROM audit_log "
        "WHERE table_name = 'brands' AND operation = 'CREATE' "
        "ORDER BY performed_at DESC LIMIT 1"
    )
    assert row is not None, "expected a CREATE audit entry"
    assert row["operation"] == "CREATE"
    assert row["old_data"] is None
    assert row["new_data"]["brand_name"] == "Audit Create Brand"
    assert row["performed_by"] == "adminuser"


async def test_import_update_writes_audit_entry(client, admin_headers, conn):
    brand_id = await conn.fetchval(
        "INSERT INTO brands (brand_name) VALUES ('Audit Update Before') RETURNING brand_id"
    )
    data = _make_xlsx({"Brands": [{"ID": str(brand_id), "Name": "Audit Update After"}]})
    res = await client.post(
        "/admin/import/xlsx",
        headers=admin_headers,
        files={"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200

    row = await conn.fetchrow(
        "SELECT operation, old_data, new_data, performed_by FROM audit_log "
        "WHERE table_name = 'brands' AND record_id = $1 AND operation = 'UPDATE'",
        brand_id,
    )
    assert row is not None, "expected an UPDATE audit entry"
    assert row["old_data"]["brand_name"] == "Audit Update Before"
    assert row["new_data"]["brand_name"] == "Audit Update After"
    assert row["performed_by"] == "adminuser"
